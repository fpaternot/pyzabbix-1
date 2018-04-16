# -*- coding: utf-8 -*-
"""
Created on Mon Jul 10 09:28:01 2017

@author: david.goncalves
"""

# do the import 
from zabbix_api import ZabbixAPI
import sys
reload(sys)
sys.setdefaultencoding('utf8')


class REPORT:
    
    def __init__(self, groupid):
        # inicialize the conection with zabbix server using the API
        self.zapi = ZabbixAPI(server="http://URL.Zabbix", timeout=120)
        self.zapi.login(user="UserName", password="PassWord")
        self.groupid = groupid
        
    def Get_Group_Name(self):
        group_name = self.zapi.hostgroup.get({"output":["name"],
                                              "groupids":self.groupid})
        
        return group_name
    
    # get the hosts of a group
    def Get_Hosts(self):
        hosts = self.zapi.host.get({"output":"extend", 
                           "groupids":self.groupid, 
                           "monitored_hosts":True})
        return hosts
        
    # define the function to translate timestamp to human date  
    def Get_Time_Human(self, clock):
        import datetime
        return datetime.datetime.fromtimestamp(
            int(clock)).strftime('%d-%m-%Y %H:%M:%S')
            
    def Incial_Load(self, hosts):
        
        # Configuring the variable that will store all the data,
        #  until put in the xls filename
        self.EVENTOS_SERVIDORES = []
        for h in hosts:
            self.data = []
            events= self.zapi.event.get({"output":"extend", 
                                    "hostids": h.get("hostid"),
                                    "time_from":"1496275278",
                                    "time_till":"1498867158",
                                    "sortfield":"clock",
                                    "limit":100, # Set limit to teste de script
                                    })
        
            for e in events:        
                trigger  = self.zapi.trigger.get({"output":"extend", 
                                         "triggerids":e.get("objectid"),
                                         "expandDescription":True,
                                         })
                evnt = (self.Get_Time_Human(e["clock"]),
                                   h["name"],
                                   trigger[0]["priority"],
                                   trigger[0]["description"].encode('utf-8'),
                                   trigger[0]["value"],
                                   e.get("acknowledged")
                        )
                self.data.append(evnt)
        
            self.EVENTOS_SERVIDORES.append(self.data)
        return self.EVENTOS_SERVIDORES
     
    # define the function how to write the data in the xls file       
    def Put_In_XLS(self, data):
        from openpyxl import Workbook
        wb = Workbook()
        FILENAME = 'Eventos {}.xlsx'.format(str(self.Get_Group_Name()[0]["name"]))
        ws1 = wb.active
        ws1.title = "Eventos do Mes"
        HEADER = ("Data do Evento",
                  "HostName",
                  "Prioridade",
                  "Descrição",
                  "Status",
                  "ACK")
                  
        for a in data:
            try:
                try:
                    ws1 = wb.create_sheet(title=a[0][1][:30])
                    
                except:
                    ws1 = wb.create_sheet(title="NULL")
                
                ws1.append(HEADER)
                
                for l in a:
                    try:
                        ws1.append(l)
                        
                    except Exception as error:
                        print "Erro ao adicionar dados"
                        print error
    
            except Exception as error:
                print "Erro ao criar/adicionar guia"
                print error
                
        wb.save(filename = FILENAME)

    def START(self):
        print "Get hosts"
        hosts = self.Get_Hosts()
        
        grupo = self.Get_Group_Name()
        print grupo[0]["name"]
        
        print "Get events"
        data = self.Incial_Load(hosts)
        
        print "Put events in xls file"
        self.Put_In_XLS(data)


    
    
# Start the report generation
if __name__ == "__main__":
    try:
            report = REPORT(21)
            print "Start Process"
            report.START()
            print "Executado com sucesso"
            
    except Exception as error:
            print "Erro ao criar o relatorio"
            print error
